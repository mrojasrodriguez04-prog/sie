from flask import jsonify,Flask, request, render_template, redirect, session, url_for,flash
from werkzeug.security import check_password_hash, generate_password_hash
from models import db,Sector,TipoFlota, Competidor,ProductoServicio,HistoricoVentas, CategoriaProductoServicio, Empresario, Subsector, Ciudad, Empresa, Usuario, Sede,RedSocial, ProcesoEmpresarial, Cargo, Stakeholder,CanalVenta,Problematica,Infraestructura,SoftwareUsado
from datetime import timedelta
from datetime import datetime
from config import Config
from sqlalchemy import distinct
from collections import Counter
from io import BytesIO
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.units import inch
from flask import send_file
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.units import cm
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config.from_object(Config)

app.secret_key = Config.SECRET_KEY
db.init_app(app)

TIPO_PERSONA = ['Natural','Jurídica']
ROL_EMPRESARIO = ['Propietario','Representante Legal','Otro']
app.permanent_session_lifetime = timedelta(days=7)
@app.route("/")
def inicio():
    return render_template("login.html")


# LOGIN
@app.route("/validar_login", methods=["POST"])
def validar_login():

    usuario = Usuario.query.filter_by(email=request.form["email"]).first()

    if usuario and check_password_hash(usuario.password, request.form["password"]):

        session["usuario_id"] = usuario.id_usuario
        session["tipo_usuario"] = usuario.tipo_usuario
        session["nombre"] = usuario.nombres
        recordar = request.form.get('recordar')

        if usuario.tipo_usuario == "admin":
            return redirect(url_for("panel"))

        if recordar:
            session.permanent = True  # dura varios días
        else:
            session.permanent = False  # se borra al cerrar navegador

        return redirect(url_for("panel"))

    else:
        flash("Datos incorrectos")
        return redirect(url_for("inicio"))


@app.route("/panel")
def panel():
    ahora = datetime.now()

    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0)
    inicio_dia = ahora.replace(hour=0, minute=0, second=0)

    if "usuario_id" not in session:
        return redirect(url_for("inicio"))
        
    total_empresarios = Empresario.query.count()
    total_empresas = Empresa.query.count()
    total_usuarios = Usuario.query.count()

    nuevos_empresarios = Empresario.query.filter(
        Empresario.fecha_registro >= inicio_mes
    ).count()

    nuevas_empresas = Empresa.query.filter(
        Empresa.fecha_registro >= inicio_mes
    ).count()

    nuevos_usuarios = Usuario.query.filter(
        Usuario.fecha_registro >= inicio_dia
    ).count()


    return render_template(
        "Panel.html", tipo=session["tipo_usuario"], nombre=session["nombre"],
        total_empresarios=total_empresarios,
        total_empresas=total_empresas,
        total_usuarios=total_usuarios,
        nuevos_empresarios=nuevos_empresarios,
        nuevas_empresas=nuevas_empresas,
        nuevos_usuarios=nuevos_usuarios)


#Usuario Registro
@app.route("/registro_usuario")
def registro_usuario():

    if session.get("tipo_usuario") != "admin":
        return redirect(url_for("panel"))

    return render_template("registro.html")


@app.route("/guardar_usuario", methods=["POST"])
def guardar_usuario():

    if session.get("tipo_usuario") != "admin":
        return "Acceso denegado"
    
    email = request.form["email"]

    usuario_existente = Usuario.query.filter_by(email=email).first()

    if usuario_existente:
        flash("⚠️ El correo ya está registrado")
        return redirect(url_for("registro_usuario"))

    password = generate_password_hash(request.form["password"])
    

    nuevo = Usuario(
        nombres=request.form["nombres"],
        apellidos=request.form["apellidos"],
        tipo_doc=request.form["tipo_doc"],
        numero_doc=request.form["numero_doc"],
        email=request.form["email"],
        password=password,
        tipo_usuario=request.form["tipo_usuario"]
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Usuario creado correctamente")
    return redirect(url_for("panel"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("inicio"))



#Empresarios y Empresas

@app.route("/registrar_empresario")
def registrar_empresario():

    return render_template(
        "registrar_empresario.html",
        tipos=TIPO_PERSONA,
        roles=ROL_EMPRESARIO
    )

@app.route("/guardar_empresario", methods=["POST"])
def guardar_empresario():

    nombres = request.form["nombres"]
    apellidos = request.form["apellidos"]
    correo = request.form["correo"]
    tipo = request.form["tipo"]
    rol = request.form["rol"]
    
    correo_existente = Empresario.query.filter_by(correo_personal=correo).first()

    if correo_existente:
        flash("⚠️ Este correo ya está registrado")
        return redirect(url_for("registrar_empresario"))

    nuevo = Empresario(
        nombres_completos=nombres,
        apellidos=apellidos,
        correo_personal=correo,
        tipo_persona=tipo,
        rol_empresario=rol
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Empresario registrado correctamente")

    return redirect(url_for("listar_empresarios"))

@app.route("/empresarios")
def listar_empresarios():

    empresarios = Empresario.query.all()

    return render_template("listar_empresarios.html", empresarios=empresarios)

@app.route("/eliminar_empresario/<int:id>")

def eliminar_empresario(id):

    empresario = Empresario.query.get(id)

    db.session.delete(empresario)
    db.session.commit()

    flash("Empresario eliminado correctamente")

    return redirect(url_for("listar_empresarios"))




@app.route("/empresa/registrar/<int:id>")
def registrar_empresa(id):
    
    empresario = Empresario.query.filter_by(id_empresario=id).first()

    if not empresario:
        flash("Empresario no encontrado")
        return redirect(url_for("listar_empresarios"))
    
    subsectores = Subsector.query.all()
    ciudades = Ciudad.query.all()

    return render_template(
        "registrar_empresa.html",
        empresario=empresario,
        subsectores=subsectores,
        ciudades=ciudades
    )

@app.route("/empresa/guardar/<int:id_empresario>", methods=["POST"])
def guardar_empresa(id_empresario):

    nueva_empresa = Empresa(
        id_empresario=id_empresario,
        nombre_empresa=request.form["nombre_empresa"],
        tipo_oferta=request.form["tipo_oferta"],
        actividad_economica=request.form["actividad"],
        tipo_persona_juridica=request.form["tipo_persona"],
        tamano_empresa=request.form["tamano"],
        punto_venta=request.form["punto_venta"],
        direccion_comercial=request.form["direccion"],
        numero_empleados=request.form["empleados"],
        telefono_contacto=request.form["telefono"],
        correo_empresarial=request.form["correo"],
        sitio_web=request.form["web"],
        id_subsector=request.form["subsector"],
        id_ciudad=request.form["ciudad"]
    )

    db.session.add(nueva_empresa)
    db.session.commit()
    nueva_flota = TipoFlota(
        id_empresa=nueva_empresa.id_empresa,
        tipo_flota=request.form["tipo_flota"]
    )

    db.session.add(nueva_flota)
    db.session.commit()

    flash("Empresa registrada correctamente")

    return redirect(url_for("empresas_por_empresario", id=id_empresario))


@app.route("/empresario/<int:id>/empresas")
def empresas_por_empresario(id):

    empresario = Empresario.query.get_or_404(id)
    empresas = Empresa.query.filter_by(id_empresario=id).all()

    return render_template(
        "empresas_por_empresario.html",
        empresas=empresas,
        empresario=empresario
    )

    
@app.route("/empresa/eliminar/<int:id>")
def eliminar_empresa(id):

    empresa = Empresa.query.get(id)
    id_empresario = empresa.id_empresario

    db.session.delete(empresa)
    db.session.commit()

    flash("Empresa eliminada correctamente")

    return redirect(url_for(
        "empresas_por_empresario",
        id=id_empresario
    ))



#Sedes
@app.route("/sede/nueva/<int:id_empresa>")
def nueva_sede(id_empresa):

    empresa = Empresa.query.get_or_404(id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)
    return render_template("registrar_sede.html", empresa=empresa, empresario=empresario)

@app.route("/guardar_sede/<int:id_empresa>", methods=["POST"])
def guardar_sede(id_empresa):

    nueva = Sede(
        id_empresa=id_empresa,
        nombre_sede=request.form["nombre_sede"],
        direccion=request.form["direccion"]
    )

    db.session.add(nueva)
    db.session.commit()

    flash("Sede registrada correctamente")

    return redirect(url_for("listar_sedes", id_empresa=id_empresa))

@app.route("/empresa/<int:id_empresa>/sedes")
def listar_sedes(id_empresa):

    empresa = Empresa.query.get_or_404(id_empresa)
    sedes = Sede.query.filter_by(id_empresa=id_empresa).all()
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template(
        "listar_sedes.html",
        sedes=sedes,
        empresa=empresa,
        empresario=empresario
    )

@app.route("/sede/eliminar/<int:id>")
def eliminar_sede(id):

    sede = Sede.query.get(id)
    id_empresa = sede.id_empresa

    db.session.delete(sede)
    db.session.commit()

    flash("Sede eliminada correctamente")

    return redirect(url_for("listar_sedes", id_empresa=id_empresa))

#redes_sociales
@app.route("/red/nueva/<int:id_empresa>")
def nueva_red(id_empresa):

    empresa = Empresa.query.get_or_404(id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)
    return render_template("registrar_red.html", empresa=empresa, empresario=empresario)

@app.route("/guardar_red/<int:id_empresa>", methods=["POST"])
def guardar_red(id_empresa):

    nueva = RedSocial(
        id_empresa=id_empresa,
        tipo_red=request.form["tipo_red"],
        url_red=request.form["url_red"]
    )

    db.session.add(nueva)
    db.session.commit()

    flash("Red social registrada correctamente")

    return redirect(url_for("listar_redes", id_empresa=id_empresa))

@app.route("/empresa/<int:id_empresa>/redes")
def listar_redes(id_empresa):

    empresa = Empresa.query.get_or_404(id_empresa)
    
    redes = RedSocial.query.filter_by(id_empresa=id_empresa).all()  
    empresario = Empresario.query.get(empresa.id_empresario)

    

    return render_template(
        "listar_redes.html",
        redes=redes,
        empresa=empresa,
        empresario=empresario 
    )
    
@app.route("/red/eliminar/<int:id>")
def eliminar_red(id):

    red = RedSocial.query.get(id)
    id_empresa = red.id_empresa

    db.session.delete(red)
    db.session.commit()

    flash("Red social eliminada correctamente")

    return redirect(url_for("listar_redes", id_empresa=id_empresa))
    
#procesos
@app.route("/proceso/nuevo/<int:id_empresa>")
def nuevo_proceso(id_empresa):

    empresa = Empresa.query.get(id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template(
        "registro_proceso.html",
        empresa=empresa, empresario=empresario
    )


@app.route("/guardar_proceso/<int:id_empresa>", methods=["POST"])
def guardar_proceso(id_empresa):

    nuevo = ProcesoEmpresarial(
        id_empresa=id_empresa,
        subproceso_area=request.form["subproceso_area"],
        tipo_proceso=request.form["tipo_proceso"]
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Proceso registrado correctamente")

    return redirect(url_for("listar_procesos", id_empresa=id_empresa))

@app.route("/empresa/<int:id_empresa>/procesos")
def listar_procesos(id_empresa):

    empresa = Empresa.query.get(id_empresa)
    procesos = ProcesoEmpresarial.query.filter_by(id_empresa=id_empresa).all()
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template(
        "listar_procesos.html",
        empresa=empresa,
        procesos=procesos, empresario=empresario
    )

@app.route("/proceso/eliminar/<int:id>")
def eliminar_proceso(id):

    proceso = ProcesoEmpresarial.query.get(id)
    id_empresa = proceso.id_empresa

    db.session.delete(proceso)
    db.session.commit()

    flash("Proceso eliminado")

    return redirect(url_for("listar_procesos", id_empresa=id_empresa))

#cargos

@app.route("/cargo/nuevo/<int:id_proceso>")
def nuevo_cargo(id_proceso):

    proceso = ProcesoEmpresarial.query.get_or_404(id_proceso)
    empresa = Empresa.query.get(proceso.id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template(
        "registro_cargo.html",
        proceso=proceso,
        empresa=empresa, empresario=empresario
    )

@app.route("/guardar_cargo/<int:id_proceso>", methods=["POST"])
def guardar_cargo(id_proceso):

    nuevo = Cargo(
        id_proceso=id_proceso,
        nombre_cargo=request.form["nombre_cargo"],
        cantidad_empleados=request.form["cantidad"]
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Cargo registrado correctamente")

    return redirect(url_for("listar_cargos", id_proceso=id_proceso))

@app.route("/proceso/<int:id_proceso>/cargos")
def listar_cargos(id_proceso):

    proceso = ProcesoEmpresarial.query.get_or_404(id_proceso)
    empresa = Empresa.query.get(proceso.id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)

    cargos = Cargo.query.filter_by(id_proceso=id_proceso).all()

    return render_template(
        "listar_cargos.html",
        proceso=proceso,
        empresa=empresa,
        cargos=cargos, empresario=empresario
    )

@app.route("/cargo/eliminar/<int:id>")
def eliminar_cargo(id):

    cargo = Cargo.query.get(id)

    if cargo:
        id_proceso = cargo.id_proceso
        db.session.delete(cargo)
        db.session.commit()
        flash("Cargo eliminado correctamente")
        return redirect(url_for("listar_cargos", id_proceso=id_proceso))

    flash("Cargo no encontrado")
    return redirect(url_for("panel"))

#Stakeholders
@app.route("/stakeholder/nuevo/<int:id_empresa>")
def nuevo_stakeholder(id_empresa):
    empresa = Empresa.query.get(id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template(
        "registro_stakeholder.html",
        empresa=empresa,
        empresario=empresario
    )

@app.route("/guardar_stakeholder/<int:id_empresa>", methods=["POST"])
def guardar_stakeholder(id_empresa):
    nuevo = Stakeholder(
        id_empresa=id_empresa,
        nombre_area=request.form["nombre_area"]
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Stakeholder registrado correctamente")

    return redirect(url_for('listar_stakeholders', id=id_empresa))

@app.route("/empresa/<int:id>/stakeholders")
def listar_stakeholders(id):
    empresa = Empresa.query.get(id)
    empresario = Empresario.query.get(empresa.id_empresario)

    stakeholders = Stakeholder.query.filter_by(id_empresa=id).all()

    return render_template(
        "listar_stakeholder.html",
        empresa=empresa,
        empresario=empresario,
        stakeholders=stakeholders
    )

@app.route("/stakeholder/eliminar/<int:id>")
def eliminar_stakeholder(id):
    s = Stakeholder.query.get(id)

    if s:
        db.session.delete(s)
        db.session.commit()

    return redirect(request.referrer)

#Canal de venta
@app.route("/canal/nuevo/<int:id_empresa>")
def nuevo_canal(id_empresa):
    empresa = Empresa.query.get(id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template(
        "registro_canal.html",
        empresa=empresa,
        empresario=empresario
    )

@app.route("/guardar_canal/<int:id_empresa>", methods=["POST"])
def guardar_canal(id_empresa):
    nuevo = CanalVenta(
        id_empresa=id_empresa,
        canal=request.form["canal"]
    )

    db.session.add(nuevo)
    db.session.commit()
    flash("Canal Registrado correctamente")
    return redirect(url_for('listar_canales', id=id_empresa))
    
@app.route("/empresa/<int:id>/canales")
def listar_canales(id):
    empresa = Empresa.query.get(id)
    empresario = Empresario.query.get(empresa.id_empresario)

    canales = CanalVenta.query.filter_by(id_empresa=id).all()

    return render_template(
        "listar_canal.html",
        empresa=empresa,
        empresario=empresario,
        canales=canales
    )
@app.route("/canal/eliminar/<int:id>")
def eliminar_canal(id):
    c = CanalVenta.query.get(id)

    if c:
        db.session.delete(c)
        db.session.commit()

    return redirect(request.referrer)
#problematca-procesos
@app.route("/problema/nuevo/<int:id_proceso>")
def nuevo_problema(id_proceso):
    proceso = ProcesoEmpresarial.query.get_or_404(id_proceso)
    empresa = Empresa.query.get(proceso.id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template(
        "registro_problema.html",
        proceso=proceso,
        empresa=empresa, empresario=empresario
    )
@app.route("/guardar_problema/<int:id_proceso>", methods=["POST"])
def guardar_problema(id_proceso):
    nueva = Problematica(
        id_proceso=id_proceso,
        descripcion=request.form["descripcion"]
    )

    db.session.add(nueva)
    db.session.commit()

    flash("Problemática registrada correctamente")
    return redirect(url_for("listar_problemas", id_proceso=id_proceso))
@app.route("/proceso/<int:id_proceso>/problemas")
def listar_problemas(id_proceso):
    proceso = ProcesoEmpresarial.query.get_or_404(id_proceso)
    empresa = Empresa.query.get(proceso.id_empresa)
    empresario = Empresario.query.get(empresa.id_empresario)

    problemas = Problematica.query.filter_by(id_proceso=id_proceso).all()

    return render_template(
        "listar_problemas.html",
        proceso=proceso,
        empresa=empresa,
        problemas=problemas, empresario=empresario
    )

@app.route("/problema/eliminar/<int:id>")
def eliminar_problema(id):
    problema = Problematica.query.get_or_404(id)
    id_proceso = problema.id_proceso

    db.session.delete(problema)
    db.session.commit()

    flash("Problemática eliminada")

    return redirect(url_for("listar_problemas", id_proceso=id_proceso))   
    

#Infraestructura
@app.route("/empresa/<int:id>/infraestructura")
def listar_infraestructura(id):
    empresa = Empresa.query.get_or_404(id)
    empresario = Empresario.query.get(empresa.id_empresario)

    infra = Infraestructura.query.filter_by(id_empresa=id).all()

    return render_template("listar_infraestructura.html",
                           empresa=empresa,
                           empresario=empresario,
                           infraestructuras=infra)


@app.route("/infraestructura/nuevo/<int:id>")
def nuevo_infraestructura(id):
    empresa = Empresa.query.get_or_404(id)
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template("registro_infraestructura.html",
                           empresa=empresa,
                           empresario=empresario)


@app.route("/guardar_infraestructura/<int:id>", methods=["POST"])
def guardar_infraestructura(id):
    nueva = Infraestructura(
        id_empresa=id,
        tipo=request.form["tipo"]
    )

    db.session.add(nueva)
    db.session.commit()

    flash("Infraestructura registrada")
    return redirect(f"/empresa/{id}/infraestructura")


@app.route("/infraestructura/eliminar/<int:id>")
def eliminar_infraestructura(id):
    i = Infraestructura.query.get(id)
    db.session.delete(i)
    db.session.commit()

    flash("Eliminado")
    return redirect(request.referrer)

#sofware

@app.route("/empresa/<int:id>/software")
def listar_software(id):
    empresa = Empresa.query.get_or_404(id)
    empresario = Empresario.query.get(empresa.id_empresario)

    software = SoftwareUsado.query.filter_by(id_empresa=id).all()

    return render_template("listar_software.html",
                           empresa=empresa,
                           empresario=empresario,
                           softwares=software)


@app.route("/software/nuevo/<int:id>")
def nuevo_software(id):
    empresa = Empresa.query.get_or_404(id)
    empresario = Empresario.query.get(empresa.id_empresario)

    return render_template("registro_software.html",
                           empresa=empresa,
                           empresario=empresario)


@app.route("/guardar_software/<int:id>", methods=["POST"])
def guardar_software(id):

    usa = True if request.form.get("usa_software") == "on" else False

    nuevo = SoftwareUsado(
        id_empresa=id,
        usa_software=usa,
        nombre_software=request.form["nombre"]
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Software registrado")
    return redirect(f"/empresa/{id}/software")


@app.route("/software/eliminar/<int:id>")
def eliminar_software(id):
    s = SoftwareUsado.query.get(id)
    db.session.delete(s)
    db.session.commit()

    flash("Eliminado")
    return redirect(request.referrer)


@app.route("/categorias")
def categorias():

    if session.get("tipo_usuario") != "admin":
        return redirect(url_for("panel"))

    categorias = CategoriaProductoServicio.query.all()

    return render_template(
        "categorias.html",
        categorias=categorias
    )


@app.route("/categoria/nueva")
def nueva_categoria():

    if session.get("tipo_usuario") != "admin":
        return redirect(url_for("panel"))

    subsectores = Subsector.query.all()

    return render_template(
        "nueva_categoria.html",
        subsectores=subsectores
    )


@app.route("/guardar_categoria", methods=["POST"])
def guardar_categoria():

    if session.get("tipo_usuario") != "admin":
        return redirect(url_for("panel"))

    nueva = CategoriaProductoServicio(
        id_subsector=request.form["subsector"],
        nombre_categoria=request.form["categoria"]
    )

    db.session.add(nueva)
    db.session.commit()

    flash("Categoría registrada correctamente")

    return redirect(url_for("categorias"))

@app.route("/categoria/eliminar/<int:id>")
def eliminar_categoria(id):

    if session.get("tipo_usuario") != "admin":
        return redirect(url_for("panel"))

    categoria = CategoriaProductoServicio.query.get(id)

    db.session.delete(categoria)
    db.session.commit()

    flash("Categoría eliminada")

    return redirect(url_for("categorias"))

@app.route("/empresa/<int:id>/productos")
def listar_productos(id):

    empresa = Empresa.query.get(id)

    productos = ProductoServicio.query.filter_by(
        id_empresa=id
    ).all()

    empresario = Empresario.query.get(
        empresa.id_empresario
    )

    return render_template(
        "listar_productos.html",
        empresa=empresa,
        empresario=empresario,
        productos=productos
    )

@app.route("/producto/nuevo/<int:id_empresa>")
def nuevo_producto(id_empresa):

    empresa = Empresa.query.get(id_empresa)

    empresario = Empresario.query.get(
        empresa.id_empresario
    )

    categorias = CategoriaProductoServicio.query.filter_by(
        id_subsector=empresa.id_subsector
    ).all()

    return render_template(
        "nuevo_producto.html",
        empresa=empresa,
        empresario=empresario,
        categorias=categorias
    )

@app.route("/guardar_producto/<int:id_empresa>", methods=["POST"])
def guardar_producto(id_empresa):

    nuevo = ProductoServicio(

        id_empresa=id_empresa,

        id_categoria=request.form["categoria"],

        nombre_producto=request.form["nombre"],

        precio=request.form["precio"],

        fecha_precio=request.form["fecha"]
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Producto registrado")

    return redirect(
        url_for(
            "listar_productos",
            id=id_empresa
        )
    )

@app.route("/producto/eliminar/<int:id>")
def eliminar_producto(id):

    producto = ProductoServicio.query.get(id)

    id_empresa = producto.id_empresa

    db.session.delete(producto)
    db.session.commit()

    flash("Producto eliminado")

    return redirect(
        url_for(
            "listar_productos",
            id=id_empresa
        )
    )

@app.route("/producto/<int:id>/ventas")
def listar_ventas(id):

    producto = ProductoServicio.query.get(id)

    empresa = Empresa.query.get(
        producto.id_empresa
    )

    empresario = Empresario.query.get(
        empresa.id_empresario
    )

    ventas = HistoricoVentas.query.filter_by(
        id_producto=id
    ).all()

    return render_template(
        "listar_ventas.html",
        producto=producto,
        empresa=empresa,
        empresario=empresario,
        ventas=ventas
    )
@app.route("/venta/nueva/<int:id_producto>")
def nueva_venta(id_producto):

    producto = ProductoServicio.query.get(id_producto)

    empresa = Empresa.query.get(
        producto.id_empresa
    )

    empresario = Empresario.query.get(
        empresa.id_empresario
    )

    return render_template(
        "nueva_venta.html",
        producto=producto,
        empresa=empresa,
        empresario=empresario
    )

@app.route("/guardar_venta/<int:id_producto>", methods=["POST"])
def guardar_venta(id_producto):

    nueva = HistoricoVentas(

        id_producto=id_producto,

        fecha_inicio=request.form["fecha_inicio"],

        fecha_fin=request.form["fecha_fin"],

        unidades_vendidas=request.form["unidades"],

        valor_ventas=request.form["valor"]
    )

    db.session.add(nueva)
    db.session.commit()

    flash("Histórico registrado")

    return redirect(
        url_for(
            "listar_ventas",
            id=id_producto
        )
    )

@app.route("/venta/eliminar/<int:id>")
def eliminar_venta(id):

    venta = HistoricoVentas.query.get(id)

    id_producto = venta.id_producto

    db.session.delete(venta)
    db.session.commit()

    flash("Histórico eliminado")

    return redirect(
        url_for(
            "listar_ventas",
            id=id_producto
        )
    )

@app.route("/empresa/<int:id>/competidores")
def listar_competidores(id):

    empresa = Empresa.query.get(id)

    empresario = Empresario.query.get(
        empresa.id_empresario
    )

    competidores = Competidor.query.filter_by(
        id_empresa=id
    ).all()

    return render_template(
        "listar_competidores.html",
        empresa=empresa,
        empresario=empresario,
        competidores=competidores
    )
@app.route("/competidor/nuevo/<int:id_empresa>")
def nuevo_competidor(id_empresa):

    empresa = Empresa.query.get(id_empresa)

    empresario = Empresario.query.get(
        empresa.id_empresario
    )

    empresas_subsector = Empresa.query.filter(
        Empresa.id_subsector == empresa.id_subsector,
        Empresa.id_empresa != empresa.id_empresa
    ).all()

    return render_template(
        "nuevo_competidor.html",
        empresa=empresa,
        empresario=empresario,
        empresas_subsector=empresas_subsector
    )
@app.route("/guardar_competidor/<int:id_empresa>", methods=["POST"])
def guardar_competidor(id_empresa):

    nuevo = Competidor(

        id_empresa=id_empresa,

        id_empresa_competidora=request.form[
            "empresa_competidora"
        ]
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Competidor agregado")

    return redirect(
        url_for(
            "listar_competidores",
            id=id_empresa
        )
    )
@app.route("/competidor/eliminar/<int:id>")
def eliminar_competidor(id):

    competidor = Competidor.query.get(id)

    id_empresa = competidor.id_empresa

    db.session.delete(competidor)
    db.session.commit()

    flash("Competidor eliminado")

    return redirect(
        url_for(
            "listar_competidores",
            id=id_empresa
        )
    )

@app.route("/consultas")
def modulo_consultas():

    return render_template(
        "consultas.html"
    )

@app.route("/consulta/ventas")
def consulta_ventas():

    sectores = Sector.query.all()

    return render_template(
        "consulta_ventas.html",
        sectores=sectores
    )


@app.route("/get_subsectores/<int:id_sector>")
def get_subsectores(id_sector):

    subsectores = Subsector.query.filter_by(
        id_sector=id_sector
    ).all()

    return jsonify([
        {
            "id": s.id_subsector,
            "nombre": s.nombre_subsector
        }
        for s in subsectores
    ])


@app.route("/get_categorias/<int:id_subsector>")
def get_categorias(id_subsector):

    categorias = CategoriaProductoServicio.query.filter_by(
        id_subsector=id_subsector
    ).all()

    return jsonify([
        {
            "id": c.id_categoria,
            "nombre": c.nombre_categoria
        }
        for c in categorias
    ])

@app.route("/get_productos/<int:id_categoria>")
def get_productos(id_categoria):

    productos = db.session.query(
        distinct(
            ProductoServicio.nombre_producto
        )
    ).filter(
        ProductoServicio.id_categoria == id_categoria
    ).all()

    return jsonify([
        p[0]
        for p in productos
    ])




from collections import defaultdict
@app.route("/consulta/resultados", methods=["POST"])
def resultado_consulta():

    from datetime import datetime

    id_categoria = request.form["categoria"]
    nombre_producto = request.form["producto"]
    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    productos = ProductoServicio.query.filter(
        ProductoServicio.id_categoria == id_categoria,
        ProductoServicio.nombre_producto == nombre_producto
    ).all()

    resumen = []

    datasets_unidades = []
    datasets_valor = []

    colores = [
        "blue",
        "green",
        "red",
        "purple",
        "orange",
        "brown",
        "black",
        "pink",
        "cyan",
        "gray"
    ]

    fecha_inicio_label = datetime.strptime(
        fecha_inicio,
        "%Y-%m-%d"
    ).strftime("%d/%m/%y")

    fecha_fin_label = datetime.strptime(
        fecha_fin,
        "%Y-%m-%d"
    ).strftime("%d/%m/%y")

    labels = [
        fecha_inicio_label,
        fecha_fin_label
    ]

    for i, producto in enumerate(productos):

        empresa = producto.empresa

        ventas = HistoricoVentas.query.filter(
            HistoricoVentas.id_producto == producto.id_producto,
            HistoricoVentas.fecha_inicio >= fecha_inicio,
            HistoricoVentas.fecha_fin <= fecha_fin
        ).order_by(
            HistoricoVentas.fecha_inicio
        ).all()

        if not ventas:
            continue

        data_unidades = []
        data_valor = []

        cantidad_ventas = len(ventas)

        for pos, venta in enumerate(ventas):
            if cantidad_ventas == 1:
                data_unidades.append({
                    "x": 0,
                    "y": venta.unidades_vendidas
                })
                    
                data_unidades.append({
                    "x": 100,
                    "y": venta.unidades_vendidas
                })
                
                data_valor.append({
                    "x": 0,
                    "y": float(venta.valor_ventas)
                })
                
                data_valor.append({
                    "x": 100,
                    "y": float(venta.valor_ventas)
                })
                
                continue

            elif cantidad_ventas == 2:

                if pos == 0:
                    x = 0
                else:
                    x = 100

            else:

                x = round(
                    (pos * 100) /
                    (cantidad_ventas - 1),
                    2
                )
               
            data_unidades.append({

                "x": x,

                "y": venta.unidades_vendidas

            })

            data_valor.append({

                "x": x,

                "y": float(venta.valor_ventas)

            })

            resumen.append({

                "fecha":
                venta.fecha_inicio.strftime("%m/%Y"),

                "empresa":
                empresa.nombre_empresa,

                "unidades":
                venta.unidades_vendidas,

                "valor":
                float(venta.valor_ventas),

                "color":
                colores[i % len(colores)]

            })

        datasets_unidades.append({

            "label":
            empresa.nombre_empresa,

            "data":
            data_unidades,

            "borderColor":
            colores[i % len(colores)],

            "backgroundColor":
            colores[i % len(colores)],

            "fill":
            False,

            "tension":
            0.4,

            "pointRadius":
            5,

            

            "borderWidth":
            3

        })

        datasets_valor.append({

            "label":
            empresa.nombre_empresa,

            "data":
            data_valor,

            "borderColor":
            colores[i % len(colores)],

            "backgroundColor":
            colores[i % len(colores)],

            "fill":
            False,

            "tension":
            0.4,

            "pointRadius":
            5,

            "borderWidth":
            3

        })

    return render_template(

        "resultado_consulta.html",

        producto=nombre_producto,

        labels=labels,

        resumen=resumen,

        datasets_unidades=datasets_unidades,

        datasets_valor=datasets_valor

    )


@app.route("/consulta/problematicas")
def consulta_problematicas():
    return render_template(
        "consulta_problematicas.html"
        )
@app.route("/consulta/problematicas/resultados", methods=["POST"])
def resultado_problematicas():

    tipo_empresa = request.form["tipo_empresa"]
    tipo_proceso = request.form["tipo_proceso"]

    problematicas = db.session.query(
        Problematica.descripcion
    ).join(
        ProcesoEmpresarial,
        Problematica.id_proceso == ProcesoEmpresarial.id_proceso
    ).join(
        Empresa,
        ProcesoEmpresarial.id_empresa == Empresa.id_empresa
    ).filter(
        Empresa.tipo_oferta == tipo_empresa,
        ProcesoEmpresarial.tipo_proceso == tipo_proceso
    ).all()

    lista = [p[0] for p in problematicas]

    conteo = Counter(lista)

    labels = list(conteo.keys())
    valores = list(conteo.values())

    tabla = []

    for problema, cantidad in conteo.items():
        tabla.append({
            "problematica": problema,
            "cantidad": cantidad
        })

    return render_template(
        "resultado_problematicas.html",
        tipo_empresa=tipo_empresa,
        tipo_proceso=tipo_proceso,
        tabla=tabla,
        labels=labels,
        valores=valores
    )

@app.route("/consulta/empresas")
def consulta_empresas():

    sectores = Sector.query.all()

    return render_template(
        "consulta_empresas.html",
        sectores=sectores
    )

@app.route("/consulta_empresas/get_subsectores/<int:id_sector>")
def get_subsectores_empresa(id_sector):

    subsectores = Subsector.query.filter_by(
        id_sector=id_sector
    ).all()

    return jsonify([
        {
            "id": s.id_subsector,
            "nombre": s.nombre_subsector
        }
        for s in subsectores
    ])

@app.route("/consulta/buscar_empresas")
def buscar_empresas():

    tipo_empresa = request.args.get("tipo_empresa")
    sector = request.args.get("sector")
    subsector = request.args.get("subsector")
    tamano = request.args.get("tamano")

    if not tipo_empresa and not sector and not subsector and not tamano:
        return jsonify([])

    consulta = db.session.query(
        Empresa.nombre_empresa,
        ProductoServicio.nombre_producto,
        Empresa.numero_empleados
    ).join(
        ProductoServicio,
        Empresa.id_empresa == ProductoServicio.id_empresa
    ).join(
        Subsector,
        Empresa.id_subsector == Subsector.id_subsector
    ).join(
        Sector,
        Subsector.id_sector == Sector.id_sector
    )

    if tipo_empresa:
        consulta = consulta.filter(
            Empresa.tipo_oferta == tipo_empresa
        )

    if sector:
        consulta = consulta.filter(
            Sector.id_sector == sector
        )

    if subsector:
        consulta = consulta.filter(
            Empresa.id_subsector == subsector
        )

    if tamano:
        consulta = consulta.filter(
            Empresa.tamano_empresa == tamano
        )

    resultados = consulta.all()

    return jsonify([
        {
            "empresa": e.nombre_empresa,
            "producto": e.nombre_producto,
            "empleados": e.numero_empleados
        }
        for e in resultados
    ])
@app.route("/consulta/buscar_infraestructura")
def buscar_infraestructura():

    tecnologias = request.args.getlist("tecnologia")

    # Si no hay ninguna tecnología seleccionada,
    # no mostrar empresas.
    if len(tecnologias) == 0:
        return jsonify([])

    consulta = db.session.query(
        Empresa.nombre_empresa,
        Sector.nombre_sector,
        Empresa.tamano_empresa,
        Infraestructura.tipo
    ).join(
        Subsector,
        Empresa.id_subsector == Subsector.id_subsector
    ).join(
        Sector,
        Subsector.id_sector == Sector.id_sector
    ).join(
        Infraestructura,
        Empresa.id_empresa == Infraestructura.id_empresa
    ).filter(
        Infraestructura.tipo.in_(tecnologias)
    )

    resultados = consulta.all()

    return jsonify([
        {
            "empresa": r[0],
            "sector": r[1],
            "tamano": r[2],
            "tecnologia": r[3]
        }
        for r in resultados
    ])

@app.route("/consulta/exportar_excel")
def exportar_excel():

    tipo_empresa = request.args.get("tipo_empresa")
    sector = request.args.get("sector")
    subsector = request.args.get("subsector")
    tamano = request.args.get("tamano")

    consulta = db.session.query(

        Empresa.nombre_empresa,

        ProductoServicio.nombre_producto,

        Empresa.numero_empleados

    ).join(

        ProductoServicio

    ).join(

        Subsector,

        Empresa.id_subsector ==
        Subsector.id_subsector

    ).join(

        Sector,

        Subsector.id_sector ==
        Sector.id_sector

    )

    if tipo_empresa:

        consulta = consulta.filter(

            Empresa.tipo_oferta ==
            tipo_empresa

        )

    if sector:

        consulta = consulta.filter(

            Sector.id_sector ==
            sector

        )

    if subsector:

        consulta = consulta.filter(

            Subsector.id_subsector ==
            subsector

        )

    if tamano:

        consulta = consulta.filter(

            Empresa.tamano_empresa ==
            tamano

        )

    datos = consulta.all()

    df = pd.DataFrame(

        datos,

        columns=[

            "Empresa",

            "Producto / Servicio",

            "Número de empleados"

        ]

    )

    archivo = BytesIO()

    with pd.ExcelWriter(

        archivo,

        engine="openpyxl"

    ) as writer:

        df.to_excel(

            writer,

            index=False,

            startrow=7,

            sheet_name="Empresas"

        )

        libro = writer.book

        hoja = writer.sheets["Empresas"]


        # ---------------- LOGO ----------------

        try:

            logo = Image(
                "static/img/logo_sena.png"
            )

            logo.width = 75

            logo.height = 75

            hoja.add_image(
                logo,
                "A1"
            )

        except:

            pass


        # ---------------- TITULO ----------------

        hoja.merge_cells("B1:D1")

        hoja["B1"] = "SIE"

        hoja["B1"].font = Font(

            size=20,

            bold=True,

            color="012970"

        )

        hoja["B1"].alignment = Alignment(

            horizontal="center"

        )


        hoja.merge_cells("A3:D3")

        hoja["A3"] = "REPORTE DE EMPRESAS"

        hoja["A3"].font = Font(

            size=16,

            bold=True,

            color="012970"

        )

        hoja["A3"].alignment = Alignment(

            horizontal="center"

        )


        # ---------------- FECHA ----------------

        hoja["A5"] = "Fecha de generación"

        hoja["B5"] = datetime.now().strftime(

            "%d/%m/%Y %H:%M"

        )


        # ---------------- FILTROS ----------------

        hoja["A6"] = "Tipo empresa"

        hoja["B6"] = (

            tipo_empresa

            if tipo_empresa

            else "Todos"

        )

        hoja["C6"] = "Tamaño"

        hoja["D6"] = (

            tamano

            if tamano

            else "Todos"

        )


        # ---------------- ESTILOS ----------------

        verde = PatternFill(

            fill_type="solid",

            start_color="39A900",

            end_color="39A900"

        )

        blanco = Font(

            color="FFFFFF",

            bold=True

        )


        for celda in hoja[8]:

            celda.fill = verde

            celda.font = blanco

            celda.alignment = Alignment(

                horizontal="center"

            )


        borde = Border(

            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin")

        )


        for fila in hoja.iter_rows():

            for celda in fila:

                celda.border = borde


        # ---------------- AJUSTAR COLUMNAS ----------------

    for columna in range(1, hoja.max_column + 1):
        
        longitud = 0
        
        letra = get_column_letter(columna)
        
        for fila in range(1, hoja.max_row + 1):
            
            valor = hoja.cell(row=fila, column=columna).value
            
            if valor:
                longitud = max(
                    longitud,
                    len(str(valor))
                )
        hoja.column_dimensions[letra].width = longitud + 5


    archivo.seek(0)

    return send_file(

        archivo,

        download_name="Reporte_Empresas.xlsx",

        as_attachment=True,

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )
@app.route("/consulta/exportar_pdf")
def exportar_pdf():

    tipo_empresa = request.args.get("tipo_empresa")
    sector = request.args.get("sector")
    subsector = request.args.get("subsector")
    tamano = request.args.get("tamano")

    consulta = db.session.query(

        Empresa.nombre_empresa,

        ProductoServicio.nombre_producto,

        Empresa.numero_empleados

    ).join(

        ProductoServicio

    ).join(

        Subsector,
        Empresa.id_subsector == Subsector.id_subsector

    ).join(

        Sector,
        Subsector.id_sector == Sector.id_sector

    )

    if tipo_empresa:
        consulta = consulta.filter(
            Empresa.tipo_oferta == tipo_empresa
        )

    if sector:
        consulta = consulta.filter(
            Sector.id_sector == sector
        )

    if subsector:
        consulta = consulta.filter(
            Subsector.id_subsector == subsector
        )

    if tamano:
        consulta = consulta.filter(
            Empresa.tamano_empresa == tamano
        )

    datos = consulta.all()

    archivo = BytesIO()

    pdf = SimpleDocTemplate(
        archivo,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )

    estilos = getSampleStyleSheet()

    titulo = estilos["Heading1"]
    titulo.alignment = TA_CENTER
    titulo.textColor = colors.HexColor("#012970")

    texto = estilos["Normal"]

    elementos = []

    try:

        logo = Image(
            "static/img/logo_sena.png",
            width=2.2*cm,
            height=2.2*cm
        )

        elementos.append(logo)

    except:
        pass

    elementos.append(
        Paragraph(
            "<b>SIE</b>",
            estilos["Title"]
        )
    )

    elementos.append(
        Paragraph(
            "REPORTE DE EMPRESAS",
            titulo
        )
    )

    elementos.append(
        Paragraph(
            "Fecha de generación: "
            + datetime.now().strftime("%d/%m/%Y %H:%M"),
            texto
        )
    )

    elementos.append(
        Paragraph(
            f"<b>Tipo empresa:</b> {tipo_empresa if tipo_empresa else 'Todos'}",
            texto
        )
    )

    elementos.append(
        Paragraph(
            f"<b>Tamaño:</b> {tamano if tamano else 'Todos'}",
            texto
        )
    )

    elementos.append(
        Spacer(1,0.5*cm)
    )

    tabla = [[

        "Empresa",

        "Producto / Servicio",

        "Empleados"

    ]]

    for fila in datos:

        tabla.append([

            fila[0],

            fila[1],

            fila[2]

        ])

    t = Table(
        tabla,
        colWidths=[7*cm,7*cm,3*cm]
    )

    t.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#39A900")),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("ALIGN",(0,0),(-1,-1),"CENTER"),

            ("GRID",(0,0),(-1,-1),0.5,colors.black),

            ("BOTTOMPADDING",(0,0),(-1,0),8),

            ("BACKGROUND",(0,1),(-1,-1),colors.white)

        ])

    )

    elementos.append(t)

    pdf.build(elementos)

    archivo.seek(0)

    return send_file(

        archivo,

        download_name="Reporte_Empresas.pdf",

        as_attachment=True

    )
if __name__ == "__main__":
    app.run(debug=True)
